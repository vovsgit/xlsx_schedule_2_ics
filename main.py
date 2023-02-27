import os
import sys
from typing import List
import openpyxl.workbook
from openpyxl import Workbook


def check_args(args: List[str]) -> List[str]:
    if len(args) == 3:
        return [args[1], args[2]]
    else:
        raise ValueError("Invalid number of arguments")


def get_files(group: str) -> List[Workbook]:
    dir_path = './schedules'

    if group == 'all':
        files = [os.path.join(dir_path, f) for f in os.listdir(dir_path) if '.xlsx' in f]
        return [openpyxl.load_workbook(file) for file in files]
    else:
        files = [os.path.join(dir_path, f) for f in os.listdir(dir_path) if group in f]
        return [openpyxl.load_workbook(files[0])]

def parse_workbook(workbook: Workbook) -> List[List[str]]:
    sheet = workbook.active
    data = []

    #hidden_rows = [row_idx for row_idx, row in enumerate(sheet.rows) if row[0].hidden]

    # Удаление скрытых строк
    for row in reversed(list(sheet.rows)):
        if row[0].hidden:
            sheet.delete_rows(row[0].row)

    groups = sheet.cell(row=2, column=2).value
    speciality = sheet.cell(row=3, column=2).value

    hidden_cells = [(cell.row, cell.column) for cell in sheet._cells if cell.has_style('hidden')]

    disciplines = [sheet.cell(row=9 + i, column=2).value for i in range(0, 27, 3)]

    for row in sheet.iter_rows(min_row=1, max_col=71, values_only=True):
        data.append(row)
    return data

if __name__ == "__main__":
    args = sys.argv
    group, english_group = check_args(args)
    workbooks = get_files(group)
    parse_workbook(workbooks[0])

